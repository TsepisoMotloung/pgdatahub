"""Excel processing module with streaming and chunked reading."""

import pandas as pd
from pathlib import Path
from typing import Iterator, Dict, List, Tuple, Optional, Any
from datetime import datetime

from .utils import (
    logger, normalize_column_names, compute_file_hash,
    get_folder_path_parts
)
from .config import get_config


class ExcelProcessor:
    """Process Excel files with streaming and memory efficiency."""

    def __init__(self):
        self.config = get_config()
        self.chunk_size = self.config.chunk_size

    def get_sheet_names(self, file_path: Path) -> List[str]:
        """Get list of sheet names in an Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of sheet names
        """
        try:
            xl = pd.ExcelFile(file_path, engine='openpyxl')
            return xl.sheet_names
        except Exception as e:
            logger.error(f"Error reading sheet names from {file_path}: {e}")
            return []

    def has_sheet(self, file_path: Path, sheet_name: str) -> bool:
        """Check if a sheet exists in the Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to check

        Returns:
            True if sheet exists
        """
        return sheet_name in self.get_sheet_names(file_path)

    def read_excel_chunks(self, file_path: Path, sheet_name: str,
                          columns: List[str] = None) -> Iterator[pd.DataFrame]:
        """Read Excel file in chunks using streaming.

        Uses openpyxl in read_only mode for memory efficiency.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to read
            columns: Optional column names to use (if different from header)

        Yields:
            DataFrame chunks
        """
        try:
            # Use openpyxl with read_only=True for memory efficiency
            xl = pd.ExcelFile(file_path, engine='openpyxl')

            if sheet_name not in xl.sheet_names:
                logger.warning(f"Sheet '{sheet_name}' not found in {file_path}")
                logger.info(f"Available sheets: {xl.sheet_names}")
                return

            # Read in chunks
            chunk_iter = pd.read_excel(
                xl,
                sheet_name=sheet_name,
                chunksize=self.chunk_size,
                engine='openpyxl',
                dtype=str,  # Read all as strings initially
                keep_default_na=True
            )

            for chunk in chunk_iter:
                yield chunk

        except Exception as e:
            logger.error(f"Error reading {file_path}: {e}")
            raise

    def read_excel_streaming(self, file_path: Path, sheet_name: str) -> Iterator[pd.DataFrame]:
        """Read Excel file with true streaming for very large files.

        This method uses openpyxl's read_only mode directly for maximum
        memory efficiency with very large Excel files.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to read

        Yields:
            DataFrame chunks
        """
        try:
            from openpyxl import load_workbook

            # Load workbook in read_only mode
            wb = load_workbook(file_path, read_only=True, data_only=True)

            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in {file_path}")
                logger.info(f"Available sheets: {wb.sheetnames}")
                wb.close()
                return

            ws = wb[sheet_name]

            # Read header row
            header_row = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                header_row.append(cell.value)

            # Normalize column names
            normalized_columns = normalize_column_names(
                [str(h) if h is not None else f'col_{i}'
                 for i, h in enumerate(header_row)]
            )

            # Process rows in chunks
            chunk_rows = []
            row_count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                chunk_rows.append(row)
                row_count += 1

                if len(chunk_rows) >= self.chunk_size:
                    yield self._rows_to_dataframe(chunk_rows, normalized_columns)
                    chunk_rows = []

            # Yield remaining rows
            if chunk_rows:
                yield self._rows_to_dataframe(chunk_rows, normalized_columns)

            wb.close()

        except Exception as e:
            logger.error(f"Error streaming {file_path}: {e}")
            raise

    def _rows_to_dataframe(self, rows: List[tuple], columns: List[str]) -> pd.DataFrame:
        """Convert list of row tuples to DataFrame.

        Args:
            rows: List of row value tuples
            columns: Column names

        Returns:
            DataFrame
        """
        # Ensure we have enough columns
        if rows and len(rows[0]) > len(columns):
            # Add generic column names for extra columns
            extra_cols = [f'extra_col_{i}' for i in range(len(columns), len(rows[0]))]
            columns = columns + extra_cols

        df = pd.DataFrame(rows, columns=columns[:len(rows[0]) if rows else 0])

        # Convert data types intelligently
        df = self._infer_types(df)

        return df

    def _infer_types(self, df: pd.DataFrame) -> pd.DataFrame:
        """Infer and convert column types.

        Args:
            df: Input DataFrame

        Returns:
            DataFrame with inferred types
        """
        for col in df.columns:
            if col in ('source_file', 'load_timestamp'):
                continue

            series = df[col]

            # Try numeric conversion
            try:
                numeric_series = pd.to_numeric(series, errors='coerce')
                if numeric_series.notna().sum() / len(series) > 0.5:
                    # More than 50% numeric - use numeric
                    df[col] = numeric_series
                    continue
            except:
                pass

            # Try datetime conversion
            try:
                datetime_series = pd.to_datetime(series, errors='coerce')
                if datetime_series.notna().sum() / len(series) > 0.5:
                    df[col] = datetime_series
                    continue
            except:
                pass

            # Keep as string/text
            df[col] = series.astype(str).replace('nan', None).replace('None', None)

        return df

    def process_file(self, file_path: Path, sheet_name: str,
                     table_name: str, folder_path: str = '') -> Tuple[int, str]:
        """Process a single Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to read
            table_name: Target database table name
            folder_path: Relative folder path

        Returns:
            Tuple of (row_count, file_hash)
        """
        logger.info(f"Processing file: {file_path}")

        # Compute file hash
        file_hash = compute_file_hash(file_path)

        # Check if sheet exists
        if not self.has_sheet(file_path, sheet_name):
            logger.warning(f"Sheet '{sheet_name}' not found in {file_path}")
            return 0, file_hash

        total_rows = 0

        # Process file in chunks
        for chunk in self.read_excel_streaming(file_path, sheet_name):
            # Add metadata columns
            chunk['source_file'] = str(file_path)
            chunk['load_timestamp'] = datetime.utcnow()

            yield chunk
            total_rows += len(chunk)

        logger.info(f"Processed {total_rows} rows from {file_path}")

    def get_file_info(self, file_path: Path, sheet_name: str) -> Dict[str, Any]:
        """Get information about an Excel file without reading all data.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to check

        Returns:
            Dict with file info
        """
        info = {
            'file_path': str(file_path),
            'file_size': file_path.stat().st_size,
            'sheet_name': sheet_name,
            'sheet_exists': False,
            'row_count': 0,
            'columns': []
        }

        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)
            info['sheet_exists'] = sheet_name in wb.sheetnames

            if info['sheet_exists']:
                ws = wb[sheet_name]

                # Get column names from header
                header_row = []
                for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                    header_row.append(str(cell.value) if cell.value else '')

                info['columns'] = normalize_column_names(header_row)

                # Count rows (approximate for read_only)
                row_count = 0
                for _ in ws.iter_rows(min_row=2):
                    row_count += 1
                info['row_count'] = row_count

            wb.close()

        except Exception as e:
            logger.error(f"Error getting file info for {file_path}: {e}")

        return info

    def coalesce_duplicate_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Coalesce values from duplicate column names.

        When duplicate columns exist, uses first non-NULL value.

        Args:
            df: DataFrame with potential duplicate columns

        Returns:
            DataFrame with coalesced duplicates
        """
        # Find duplicate column names
        seen = set()
        duplicates = set()

        for col in df.columns:
            if col in seen:
                duplicates.add(col)
            seen.add(col)

        if not duplicates:
            return df

        # Coalesce duplicates
        for dup_col in duplicates:
            dup_cols = [c for c in df.columns if c == dup_col]
            if len(dup_cols) > 1:
                # Coalesce: use first non-null value
                df[dup_col] = df[dup_cols].bfill(axis=1).iloc[:, 0]
                # Drop duplicate columns (keep first)
                df = df.loc[:, ~df.columns.duplicated(keep='first')]

        return df

    def validate_file(self, file_path: Path) -> Tuple[bool, List[str]]:
        """Validate an Excel file before processing.

        Args:
            file_path: Path to Excel file

        Returns:
            Tuple of (is_valid, error_messages)
        """
        errors = []

        # Check file exists
        if not file_path.exists():
            errors.append(f"File does not exist: {file_path}")
            return False, errors

        # Check file is readable
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            wb.close()
        except Exception as e:
            errors.append(f"Cannot read Excel file: {e}")
            return False, errors

        # Check file size (warn if very large)
        file_size = file_path.stat().st_size
        if file_size > 100 * 1024 * 1024:  # 100 MB
            logger.warning(f"Large file detected ({file_size / 1024 / 1024:.1f} MB): {file_path}")

        return len(errors) == 0, errors
