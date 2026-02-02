"""Excel file processing with streaming support for large files"""
import logging
import hashlib
from pathlib import Path
from typing import Iterator, Optional, Tuple
import pandas as pd
import openpyxl
from openpyxl import load_workbook

from src.config import config

logger = logging.getLogger(__name__)


def _is_old_excel_format(file_path: Path) -> bool:
    """Check if file is in legacy Excel format (.xls)"""
    return file_path.suffix.lower() == '.xls'


def compute_file_hash(file_path: Path) -> str:
    """
    Compute SHA-256 hash of a file.
    Used for deduplication.
    """
    sha256 = hashlib.sha256()
    
    with open(file_path, 'rb') as f:
        while chunk := f.read(8192):
            sha256.update(chunk)
    
    file_hash = sha256.hexdigest()
    logger.debug(f"Computed hash for {file_path.name}: {file_hash}")
    return file_hash


def discover_excel_files(root_dir: Path) -> dict:
    """
    Discover all Excel files under root directory.
    
    Returns dict: {folder_path: [file1, file2, ...]}
    Folders are represented as tuples of path parts.
    """
    if not root_dir.exists():
        logger.error(f"Root directory does not exist: {root_dir}")
        return {}
    
    excel_extensions = {'.xlsx', '.xls', '.xlsm', '.xlsb'}
    folder_files = {}
    
    for file_path in root_dir.rglob('*'):
        if file_path.suffix.lower() in excel_extensions:
            # Get folder parts relative to root
            relative_path = file_path.relative_to(root_dir)
            folder_parts = relative_path.parent.parts
            
            if not folder_parts:
                # File directly in root, skip or use special handling
                logger.warning(f"File in root directory, skipping: {file_path.name}")
                continue
            
            if folder_parts not in folder_files:
                folder_files[folder_parts] = []
            
            folder_files[folder_parts].append(file_path)
    
    logger.info(f"Discovered {sum(len(files) for files in folder_files.values())} Excel files in {len(folder_files)} folders")
    
    for folder, files in folder_files.items():
        logger.debug(f"Folder {'/'.join(folder)}: {len(files)} files")
    
    return folder_files


def read_excel_chunked(
    file_path: Path, 
    sheet_name: str,
    chunk_size: Optional[int] = None
) -> Iterator[Tuple[pd.DataFrame, bool]]:
    """
    Read Excel file in chunks for memory efficiency.
    
    Yields tuples of (DataFrame, is_first_chunk)
    
    Supports all Excel formats: .xlsx, .xls, .xlsm, .xlsb
    Uses openpyxl for modern formats (.xlsx, .xlsm, .xlsb) 
    and xlrd for legacy .xls format.
    Includes fallback logic to handle misnamed files.
    """
    if chunk_size is None:
        chunk_size = config.chunk_size
    
    logger.info(f"Reading Excel file: {file_path.name}, sheet: {sheet_name}, chunk size: {chunk_size}")
    
    # Try openpyxl first (handles .xlsx, .xlsm, .xlsb and mislabeled files)
    try:
        yield from _read_excel_chunked_openpyxl(file_path, sheet_name, chunk_size)
    except Exception as openpyxl_error:
        logger.debug(f"openpyxl failed for {file_path.name}: {openpyxl_error}. Trying xlrd as fallback.")
        
        # Fallback to xlrd for legacy .xls files
        try:
            yield from _read_excel_chunked_xlrd(file_path, sheet_name, chunk_size)
        except Exception as xlrd_error:
            # Both readers failed
            logger.error(f"Cannot read {file_path.name} with either openpyxl or xlrd")
            logger.error(f"  openpyxl error: {openpyxl_error}")
            logger.error(f"  xlrd error: {xlrd_error}")
            raise RuntimeError(f"Failed to read {file_path.name} with any available engine") from xlrd_error


def _read_excel_chunked_openpyxl(
    file_path: Path,
    sheet_name: str,
    chunk_size: int
) -> Iterator[Tuple[pd.DataFrame, bool]]:
    """
    Read modern Excel formats (.xlsx, .xlsm, .xlsb) using openpyxl.
    """
    try:
        # Load workbook in read-only mode
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in {file_path.name}. Available: {wb.sheetnames}")
            wb.close()
            return
        
        ws = wb[sheet_name]
        
        # Read rows in chunks
        rows_iter = ws.iter_rows(values_only=True)
        
        # First row is header
        try:
            header = next(rows_iter)
            if not header or all(h is None for h in header):
                logger.warning(f"Empty or invalid header in {file_path.name}")
                wb.close()
                return
            
            # Clean None values in header
            header = [str(h) if h is not None else f'col_{i}' for i, h in enumerate(header)]
            
        except StopIteration:
            logger.warning(f"Empty sheet: {sheet_name} in {file_path.name}")
            wb.close()
            return
        
        # Read data in chunks
        is_first_chunk = True
        chunk_rows = []
        
        for row in rows_iter:
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue
            
            chunk_rows.append(row)
            
            if len(chunk_rows) >= chunk_size:
                # Yield chunk
                df = pd.DataFrame(chunk_rows, columns=header)
                yield df, is_first_chunk
                
                is_first_chunk = False
                chunk_rows = []
        
        # Yield remaining rows
        if chunk_rows:
            df = pd.DataFrame(chunk_rows, columns=header)
            yield df, is_first_chunk
        
        wb.close()
        logger.info(f"Finished reading {file_path.name}")
        
    except Exception as e:
        logger.error(f"Error reading Excel file {file_path.name}: {e}", exc_info=True)
        raise


def _read_excel_chunked_xlrd(
    file_path: Path,
    sheet_name: str,
    chunk_size: int
) -> Iterator[Tuple[pd.DataFrame, bool]]:
    """
    Read legacy Excel format (.xls) using pandas with xlrd backend.
    Falls back to chunking via pandas read_excel.
    """
    try:
        # Read entire file with pandas (xlrd handles .xls format)
        # Note: For truly large .xls files, consider converting to .xlsx
        df_full = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')
        
        if df_full.empty:
            logger.warning(f"Empty sheet: {sheet_name} in {file_path.name}")
            return
        
        # Remove completely empty rows
        df_full = df_full.dropna(how='all')
        
        # Split into chunks
        is_first_chunk = True
        for i in range(0, len(df_full), chunk_size):
            chunk_df = df_full.iloc[i:i + chunk_size].reset_index(drop=True)
            yield chunk_df, is_first_chunk
            is_first_chunk = False
        
        logger.info(f"Finished reading {file_path.name}")
        
    except Exception as e:
        logger.error(f"Error reading Excel file {file_path.name} with xlrd: {e}", exc_info=True)
        raise


def get_sheet_names(file_path: Path) -> list:
    """Get all sheet names from an Excel file (supports all formats)"""
    try:
        if _is_old_excel_format(file_path):
            # For .xls files, use pandas
            xls = pd.ExcelFile(file_path, engine='xlrd')
            sheet_names = xls.sheet_names
            return sheet_names
        else:
            # For modern formats, use openpyxl
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
    except Exception as e:
        logger.error(f"Error getting sheet names from {file_path.name}: {e}")
        return []


def validate_excel_file(file_path: Path) -> bool:
    """
    Validate that an Excel file can be opened (supports all formats).
    Returns True if valid, False otherwise.
    """
    try:
        if _is_old_excel_format(file_path):
            # Validate .xls files with pandas/xlrd
            pd.read_excel(file_path, engine='xlrd', nrows=0)
        else:
            # Validate modern formats with openpyxl
            wb = load_workbook(file_path, read_only=True)
            wb.close()
        return True
    except Exception as e:
        logger.error(f"Invalid Excel file {file_path.name}: {e}")
        return False
