"""
Integration test for full ETL workflow

NOTE: This requires a PostgreSQL database.
Set DATABASE_URL environment variable before running.

Example:
    export DATABASE_URL="postgresql://postgres:password@localhost:5432/test_pgdatahub"
    pytest tests/test_integration.py -v
"""
import pytest
import os
from pathlib import Path
from openpyxl import Workbook
from sqlalchemy import text

from src.config import config
from src.database import db
from src.etl import run
from src.revert import revert_by_file


@pytest.fixture(scope="module")
def test_database():
    """Setup and teardown test database"""
    if not os.getenv('DATABASE_URL'):
        pytest.skip("DATABASE_URL not set")
    
    # Setup
    yield
    
    # Teardown - clean up test tables
    try:
        if db.engine:
            with db.get_connection() as conn:
                conn.execute(text("DROP TABLE IF EXISTS test_folder CASCADE"))
                conn.execute(text("DROP TABLE IF EXISTS etl_imports CASCADE"))
                conn.execute(text("DROP TABLE IF EXISTS etl_schema_changes CASCADE"))
                conn.commit()
    except:
        pass


@pytest.fixture
def test_data_dir(tmp_path):
    """Create temporary test data directory with Excel files"""
    # Create folder structure
    test_folder = tmp_path / "test_folder"
    test_folder.mkdir()
    
    # Create first Excel file
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Sheet1"
    ws1.append(["ID", "Name", "Value"])
    ws1.append([1, "Alice", 100])
    ws1.append([2, "Bob", 200])
    wb1.save(test_folder / "file1.xlsx")
    
    # Create second Excel file with additional column
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Sheet1"
    ws2.append(["ID", "Name", "Value", "Category"])
    ws2.append([3, "Charlie", 300, "A"])
    ws2.append([4, "David", 400, "B"])
    wb2.save(test_folder / "file2.xlsx")
    
    return tmp_path


@pytest.mark.integration
class TestFullETLWorkflow:
    """Integration tests for full ETL workflow"""
    
    def test_basic_etl_run(self, test_database, test_data_dir):
        """Test basic ETL execution"""
        # Run ETL
        run(data_root=str(test_data_dir))
        
        # Verify table was created
        assert db.table_exists("test_folder")
        
        # Verify data was imported
        with db.get_connection() as conn:
            result = conn.execute(text("SELECT COUNT(*) FROM test_folder"))
            count = result.fetchone()[0]
            assert count == 4  # 2 rows from each file
    
    def test_schema_evolution(self, test_database, test_data_dir):
        """Test that schema evolves with new columns"""
        run(data_root=str(test_data_dir))
        
        # Check that 'category' column was added
        columns = db.get_table_columns("test_folder")
        assert 'category' in [col.lower() for col in columns.keys()]
        
        # Verify schema changes were logged
        with db.get_connection() as conn:
            result = conn.execute(text("""
                SELECT COUNT(*) 
                FROM etl_schema_changes 
                WHERE table_name = 'test_folder'
            """))
            change_count = result.fetchone()[0]
            assert change_count >= 1  # At least table creation
    
    def test_deduplication(self, test_database, test_data_dir):
        """Test that duplicate files are skipped"""
        # First run
        run(data_root=str(test_data_dir))
        
        with db.get_connection() as conn:
            result = conn.execute(text("SELECT COUNT(*) FROM test_folder"))
            count_first = result.fetchone()[0]
        
        # Second run (should skip files)
        run(data_root=str(test_data_dir))
        
        with db.get_connection() as conn:
            result = conn.execute(text("SELECT COUNT(*) FROM test_folder"))
            count_second = result.fetchone()[0]
        
        # Count should be the same (no duplicates)
        assert count_first == count_second
    
    def test_metadata_columns(self, test_database, test_data_dir):
        """Test that metadata columns are added"""
        run(data_root=str(test_data_dir))
        
        # Check for metadata columns
        columns = db.get_table_columns("test_folder")
        column_names = [col.lower() for col in columns.keys()]
        
        assert 'source_file' in column_names
        assert 'load_timestamp' in column_names
    
    def test_import_tracking(self, test_database, test_data_dir):
        """Test that imports are tracked in etl_imports"""
        run(data_root=str(test_data_dir))
        
        with db.get_connection() as conn:
            result = conn.execute(text("""
                SELECT COUNT(*) 
                FROM etl_imports 
                WHERE table_name = 'test_folder'
            """))
            import_count = result.fetchone()[0]
            
            # Should have 2 imports (one per file)
            assert import_count == 2
    
    def test_revert_import(self, test_database, test_data_dir):
        """Test reverting an import"""
        run(data_root=str(test_data_dir))
        
        # Get source file
        files = list(Path(test_data_dir).rglob("*.xlsx"))
        source_file = str(files[0])
        
        # Count before revert
        with db.get_connection() as conn:
            result = conn.execute(text("SELECT COUNT(*) FROM test_folder"))
            count_before = result.fetchone()[0]
        
        # Revert
        rows_deleted = revert_by_file("test_folder", source_file)
        assert rows_deleted > 0
        
        # Count after revert
        with db.get_connection() as conn:
            result = conn.execute(text("SELECT COUNT(*) FROM test_folder"))
            count_after = result.fetchone()[0]
        
        assert count_after == count_before - rows_deleted


@pytest.mark.integration
@pytest.mark.slow
def test_large_file_chunking(test_database, tmp_path):
    """Test chunked reading of large files"""
    # Create large Excel file
    large_folder = tmp_path / "large_test"
    large_folder.mkdir()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Value"])
    
    # Add 15,000 rows (more than default chunk size)
    for i in range(15000):
        ws.append([i, f"value_{i}"])
    
    wb.save(large_folder / "large_file.xlsx")
    
    # Run ETL with smaller chunk size
    old_chunk_size = config.chunk_size
    config.chunk_size = 5000
    
    try:
        run(data_root=str(tmp_path))
        
        # Verify all data imported
        with db.get_connection() as conn:
            result = conn.execute(text("SELECT COUNT(*) FROM large_test"))
            count = result.fetchone()[0]
            assert count == 15000
    finally:
        config.chunk_size = old_chunk_size


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
