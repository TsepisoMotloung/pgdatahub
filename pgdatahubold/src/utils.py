import re
from datetime import datetime
import pandas as pd

# reuse a cleaning behavior similar to root main.py
TR_MAPPING = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iIgGuUsSoOcC")


def clean_text(text: str) -> str:
    """Normalize text: lowercase, underscores, ASCII-like, no special chars.

    - remove extension if present
    - replace special chars with underscore
    - collapse multiple underscores
    - prefix columns if starting with digit
    """
    if text is None:
        return ""

    if "." in text and not text.startswith("."):
        text = text.split(".")[0]

    text = str(text)
    text = text.replace(".", "_")
    text = text.translate(TR_MAPPING).lower()
    text = re.sub(r"[^a-z0-9_]", "_", text)
    text = re.sub(r"_+", "_", text)
    text = text.strip("_")
    if text and text[0].isdigit():
        text = "col_" + text
    return text


def normalize_dataframe_columns(df):
    """Normalize column names and coalesce duplicate names.

    - Clean column names using `clean_text`.
    - If multiple columns clean to the same name, coalesce them by taking
      the first non-null value across the duplicates for each row.
    This prevents duplicate column labels (which cause SQL insert errors).
    """
    df = df.copy()
    df.columns = [clean_text(c) for c in df.columns]

    # Build a new dataframe with unique column names. For duplicates, coalesce
    # by taking the first non-null value across the duplicate columns.
    out = pd.DataFrame(index=df.index)
    seen = {}
    for idx, name in enumerate(df.columns):
        if name in seen:
            # already processed via first occurrence
            continue
        same = df.loc[:, df.columns == name]
        if same.shape[1] == 1:
            out[name] = same.iloc[:, 0]
        else:
            # take first non-null value across the duplicate columns
            out[name] = same.bfill(axis=1).iloc[:, 0]
        seen[name] = 1

    # Ensure columns are valid SQL identifiers and non-empty. If a cleaned name
    # is empty or doesn't match the pattern, replace it with a safe name
    # 'col_<n>'. Also enforce uniqueness with numeric suffixes when necessary.
    import re

    cols = list(out.columns)
    valid_cols = []
    used = set()
    for i, c in enumerate(cols):
        new_c = c if c else f"col_{i}"
        if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", new_c):
            new_c = f"col_{i}"
        base = new_c
        suffix = 1
        while new_c in used:
            new_c = f"{base}_{suffix}"
            suffix += 1
        used.add(new_c)
        valid_cols.append(new_c)

    out.columns = valid_cols
    return out


def utcnow_iso():
    return datetime.utcnow().isoformat(sep=" ", timespec="seconds")


def read_excel_sheet_in_chunks(file_path, sheet_name, chunksize=10000):
    """Yield DataFrame chunks from an Excel sheet using openpyxl streaming.

    - Uses openpyxl in read_only mode to avoid loading entire workbook into memory.
    - Yields pandas.DataFrame objects with the header taken from the first row.
    - chunksize controls the number of data rows per yielded DataFrame.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            # empty sheet
            return
        columns = [h if h is not None else f"col_{i}" for i, h in enumerate(header)]
        batch = []
        row_count = 0
        for row in rows:
            batch.append(row)
            row_count += 1
            if row_count % chunksize == 0:
                df = pd.DataFrame(batch, columns=columns)
                yield df
                batch = []
        if batch:
            df = pd.DataFrame(batch, columns=columns)
            yield df
    finally:
        try:
            wb.close()
        except Exception:
            pass
